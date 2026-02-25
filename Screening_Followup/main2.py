# =====================================================
# FINPLOY DESIGNATION SELECTOR (Date Filter + Memory Logic)
# =====================================================
import os, sys
import pandas as pd
from datetime import datetime, timedelta
from PyQt6.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QLabel, QPushButton, QScrollArea,
    QFrame, QLineEdit, QHBoxLayout, QCheckBox, QMessageBox, QInputDialog
)
from PyQt6.QtCore import Qt
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# =====================================================
# CONFIGURATION
# =====================================================
INPUT_FILE = r"D:\matching_harsh\Screening_Followup\input\input1.xlsx"
SELECTED_FILE = "selected_designations.xlsx"
REMOVED_FILE = "removed_designations.xlsx"
OUTPUT_FILE = r"D:\matching_harsh\Screening_Followup\output\output1.xlsx"

FINPLOY_GREEN = "#4EA647"
FINPLOY_BLUE = "#206A98"
os.makedirs(os.path.dirname(os.path.abspath(OUTPUT_FILE)), exist_ok=True)

# =====================================================
# UTILITIES
# =====================================================
def load_list(path):
    """Read Excel file → lowercase set of designations"""
    if not os.path.exists(path): return set()
    try:
        df = pd.read_excel(path, engine="openpyxl")
        col = next((c for c in df.columns if "designation" in c.lower()), None)
        if col:
            return set(df[col].dropna().astype(str).str.strip().str.lower())
    except Exception as e:
        print(f"⚠️ Could not read {path}: {e}")
    return set()

def append_to_excel(path, items):
    """Append unique items to Excel file."""
    if not items: return
    existing = load_list(path)
    combined = sorted(existing.union({x.lower() for x in items}))
    pd.DataFrame({"designation": combined}).to_excel(path, index=False)
    print(f"✅ Updated {os.path.basename(path)} ({len(items)} total).")

def color_nonkey_headers(xlsx_path, keep_cols=("role", "date")):
    """Color headers not in keep_cols."""
    wb = load_workbook(xlsx_path)
    ws = wb.active
    fill = PatternFill(start_color="CFE8E0", end_color="CFE8E0", fill_type="solid")
    for i, c in enumerate(ws[1], start=1):
        if c.value not in keep_cols:
            ws.cell(row=1, column=i).fill = fill
    wb.save(xlsx_path)

# =====================================================
# PHASE 1 – LOAD INPUT & FILTER BY DATE
# =====================================================
if not os.path.exists(INPUT_FILE):
    raise FileNotFoundError(f"{INPUT_FILE} not found.")

df = pd.read_excel(INPUT_FILE, engine="openpyxl")
df.columns = [c.strip().lower() for c in df.columns]

# Detect role/designation column dynamically
role_col = next((c for c in df.columns if c in ["role", "designation"]), None)
date_col = next((c for c in df.columns if c == "date"), None)

if not role_col or not date_col:
    raise Exception("❌ input1.xlsx must have a 'Date' column and a 'Role' or 'Designation' column.")

# Ask user for number of days
app = QApplication(sys.argv)
days, ok = QInputDialog.getInt(None, "Finploy - Date Filter",
                               "Enter number of days to include:",
                               value=30, min=1, max=365)
if not ok:
    sys.exit(0)

cutoff = datetime.now() - timedelta(days=days)

def parse_date(val):
    try:
        return datetime.strptime(str(val), "%d-%m-%Y")
    except:
        return None

df["parsed_date"] = df[date_col].apply(parse_date)
filtered_df = df[df["parsed_date"].notna() & (df["parsed_date"] >= cutoff)].copy()

if filtered_df.empty:
    QMessageBox.information(None, "Finploy", f"No candidates found in last {days} days.")
    sys.exit(0)

# =====================================================
# LOAD MEMORY FILES
# =====================================================
selected_prev = load_list(SELECTED_FILE)
removed_prev = load_list(REMOVED_FILE)

designations_all = sorted(set(filtered_df[role_col].dropna().astype(str).str.strip()))
designations_filtered = [d for d in designations_all if d.lower() not in removed_prev]
preselected = [d for d in designations_filtered if d.lower() in selected_prev]
new_designations = [d for d in designations_filtered if d.lower() not in selected_prev]

print(f"🧩 Total unique: {len(designations_all)} | Preselected: {len(preselected)} | New: {len(new_designations)}")

selected_final, removed_final = set(), set()

# =====================================================
# PHASE 2 – DESIGNATION SELECTOR GUI
# =====================================================
class Selector(QWidget):
    def __init__(self, all_desg, preselected):
        super().__init__()
        self.setWindowTitle("Finploy – Role Selector")
        self.resize(700, 700)
        layout = QVBoxLayout()
        header = QLabel(f"Select Roles (filtered by last {days} days)")
        header.setStyleSheet(f"font-size:18px;font-weight:bold;color:{FINPLOY_BLUE};")
        layout.addWidget(header)

        # Search bar
        self.search = QLineEdit()
        self.search.setPlaceholderText("Search role...")
        self.search.textChanged.connect(self.filter_list)
        layout.addWidget(self.search)

        # Select All
        select_all_btn = QPushButton("Select All")
        select_all_btn.setStyleSheet(f"background-color:{FINPLOY_BLUE};color:white;font-weight:bold;")
        select_all_btn.clicked.connect(self.select_all)
        layout.addWidget(select_all_btn)

        # Scroll Area
        self.scroll = QScrollArea()
        self.frame = QFrame()
        self.vbox = QVBoxLayout(self.frame)
        self.checks = []

        for d in all_desg:
            cb = QCheckBox(d)
            cb.setChecked(d.lower() in [p.lower() for p in preselected])
            self.vbox.addWidget(cb)
            self.checks.append(cb)

        self.scroll.setWidget(self.frame)
        self.scroll.setWidgetResizable(True)
        layout.addWidget(self.scroll)

        # Submit
        submit = QPushButton("Submit & Continue")
        submit.setStyleSheet(f"background-color:{FINPLOY_GREEN};color:white;font-weight:bold;")
        submit.clicked.connect(self.submit)
        layout.addWidget(submit)
        self.setLayout(layout)

    def select_all(self):
        for cb in self.checks:
            cb.setChecked(True)

    def filter_list(self):
        term = self.search.text().lower()
        for cb in self.checks:
            cb.setVisible(term in cb.text().lower())

    def submit(self):
        for cb in self.checks:
            if cb.isChecked():
                selected_final.add(cb.text())
            else:
                removed_final.add(cb.text())
        QMessageBox.information(self, "Finploy",
                                f"✅ Selected: {len(selected_final)} | ❌ Removed: {len(removed_final)}")
        self.close()

win = Selector(designations_filtered, preselected)
win.show()
app.exec()

# =====================================================
# PHASE 3 – UPDATE MEMORY FILES & FILTER OUTPUT
# =====================================================
append_to_excel(SELECTED_FILE, selected_final)
append_to_excel(REMOVED_FILE, removed_final)

filtered_df["role_lower"] = filtered_df[role_col].str.lower()
final_df = filtered_df[filtered_df["role_lower"].isin(load_list(SELECTED_FILE))]

final_df.to_excel(OUTPUT_FILE, index=False)
color_nonkey_headers(OUTPUT_FILE)
print(f"✅ {OUTPUT_FILE} saved ({len(final_df)} rows, last {days} days).")
QMessageBox.information(None, "Finploy", f"✅ Saved {len(final_df)} filtered rows to {OUTPUT_FILE}")
try:
    import subprocess
    print("▶️ Running main3.py ...")
    subprocess.run(["python", r"D:\matching_harsh\Screening_Followup\main3.py"], check=True)
    print("✅ main3.py executed successfully!")
except Exception as e:
    print(f"❌ Failed to run main3.py: {e}")