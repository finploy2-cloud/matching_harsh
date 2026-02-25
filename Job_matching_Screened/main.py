import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import customtkinter as ctk
from tkinter import messagebox
import os
import re
import subprocess

# =====================================================
# CONFIGURATION
# =====================================================

INPUT_FILE = r"D:\matching_harsh\Job_matching_unscreened\output\screened_candidates.xlsx"
OUTPUT_DIR = r"D:\matching_harsh\Job_matching_Screened\output"

if not os.path.exists(OUTPUT_DIR):
    os.makedirs(OUTPUT_DIR)

def get_timestamped_filename(base_name):
    return os.path.join(OUTPUT_DIR, base_name)

# =====================================================
# READ & PREPARE DATA
# =====================================================

df = pd.read_excel(INPUT_FILE, engine='openpyxl')
df.columns = [str(col).strip().lower() for col in df.columns]

employment_col = next((c for c in df.columns if "employment-detail" in c), None)
if not employment_col:
    raise Exception("❌ 'employment-detail' column not found.")

if "link" not in df.columns or "location" not in df.columns:
    raise Exception("❌ Required columns missing: 'link' or 'location'")

df["name_location"] = (
    df["link"].fillna("").astype(str).str.strip() + "_" +
    df["location"].fillna("").astype(str).str.strip()
).str.replace("__", "_").str.strip("_")

def split_designation_company(text):
    if pd.isna(text):
        return "", ""
    parts = text.split(" at ", 1)
    return (parts[0].strip(), parts[1].strip()) if len(parts) == 2 else (text.strip(), "")

df["designation"], df["company"] = zip(*df[employment_col].apply(split_designation_company))

df["designation"] = df["designation"].fillna("NA")
designations = sorted(set(df["designation"].tolist()))

# =====================================================
# GUI SETUP
# =====================================================

ctk.set_appearance_mode("Light")
ctk.set_default_color_theme("green")

app = ctk.CTk()
app.title("Finploy Technologies - Candidate Designation Selector")
app.geometry("650x780")

# =====================================================
# HEADER
# =====================================================

ctk.CTkLabel(
    app, 
    text="FINPLOY TECHNOLOGIES",
    font=ctk.CTkFont(size=28, weight="bold")
).pack(pady=(15, 5))

ctk.CTkLabel(
    app, 
    text="Select Job Designations to Keep",
    font=ctk.CTkFont(size=18)
).pack(pady=8)

ctk.CTkLabel(
    app,
    text="Tip: Sales / Relationship / Field roles are auto-selected",
    font=ctk.CTkFont(size=13)
).pack(pady=(0, 10))

# =====================================================
# SCROLL FRAME FOR DESIGNATIONS (FIXED HEIGHT)
# =====================================================

scroll_frame = ctk.CTkScrollableFrame(app, width=580, height=380)
scroll_frame.pack(pady=5, padx=20, fill="both")

# Auto-select keywords
keywords = [
    "sales", "relationship", "field", "selas", "seles", "rm", "ro",
    "sso", "sfo", "senior officer", "team leader", "asm", "bm",
    "branch manager", "bsm", "home loan"
]

pattern = re.compile(r'\b(' + '|'.join(map(re.escape, keywords)) + r')\b', re.IGNORECASE)

var_dict = {}

for desg in designations:
    var = ctk.IntVar()
    if pattern.search(desg.lower()):
        var.set(1)
    chk = ctk.CTkCheckBox(scroll_frame, text=desg, variable=var)
    chk.pack(fill="x", pady=3, padx=10)
    var_dict[desg] = var

# =====================================================
# COUNT LABEL
# =====================================================

selection_label = ctk.CTkLabel(app, text="Selected: 0", font=ctk.CTkFont(size=14))
selection_label.pack(pady=5)

def update_selection_count():
    count = sum(v.get() for v in var_dict.values())
    selection_label.configure(text=f"Selected: {count}")

for v in var_dict.values():
    v.trace_add("write", lambda *args: update_selection_count())

update_selection_count()

# =====================================================
# SELECT ALL / DESELECT ALL BUTTONS
# =====================================================

button_frame = ctk.CTkFrame(app)
button_frame.pack(pady=10)

def select_all():
    for v in var_dict.values():
        v.set(1)
    update_selection_count()

def deselect_all():
    for v in var_dict.values():
        v.set(0)
    update_selection_count()

ctk.CTkButton(button_frame, text="Select All", width=120, command=select_all).grid(row=0, column=0, padx=10)
ctk.CTkButton(button_frame, text="Deselect All", width=120, command=deselect_all).grid(row=0, column=1, padx=10)

# =====================================================
# ALWAYS-VISIBLE PROCEED BUTTON
# =====================================================

def submit():
    selected = [d for d, v in var_dict.items() if v.get() == 1]
    if not selected:
        messagebox.showwarning("No Selection", "Please select at least one designation.")
        return

    filtered_df = df[df["designation"].isin(selected)].copy()
    output_file = get_timestamped_filename("output1.xlsx")
    filtered_df.to_excel(output_file, index=False, engine='openpyxl')

    wb = load_workbook(output_file)
    ws = wb.active

    pink = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")

    for col_idx, col_name in enumerate(filtered_df.columns, start=1):
        if col_name not in ["name_location", "designation", "company"]:
            ws.cell(row=1, column=col_idx).fill = pink

    wb.save(output_file)

    messagebox.showinfo("Success", f"Filtered file saved:\n{output_file}")
    app.destroy()

proceed_btn = ctk.CTkButton(
    app,
    text="Proceed → Filter Candidates",
    width=350,
    height=50,
    fg_color="#16A34A",
    hover_color="#138A3F",
    font=ctk.CTkFont(size=17, weight="bold"),
    command=submit
)
proceed_btn.pack(pady=15)

app.mainloop()
try:
    import subprocess
    print("▶️ Running main1.py ...")
    subprocess.run(["python", r"D:\matching_harsh\Job_matching_Screened\main1.py"], check=True)
    print("✅ main1.py executed successfully!")
except Exception as e:
    print(f"❌ Failed to run main1.py: {e}")