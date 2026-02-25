# =====================================================
# FINPLOY JOB MATCHING PIPELINE (PyQt6 + Preselected Memory Logic + Split Columns)
# =====================================================
import subprocess
import os, re, sys
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import gspread
from oauth2client.service_account import ServiceAccountCredentials
from PyQt6.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QLabel, QCheckBox, QPushButton,
    QScrollArea, QMessageBox, QHBoxLayout, QLineEdit, QFrame
)
from PyQt6.QtCore import Qt

# =====================================================
# CONFIGURATION
# =====================================================
OUTPUT_DIR = r"D:\matching_harsh\Job_matching_unscreened\output"
INPUT_FILE = r"D:\matching_harsh\Job_matching_unscreened\final_input\input1.xlsx"
SELECTED_FILE = r"D:\matching_harsh\Job_matching_unscreened\selected_designations.xlsx"
REMOVED_FILE = r"D:\matching_harsh\Job_matching_unscreened\removed_designations.xlsx"
SERVICE_ACCOUNT_JSON = "service_account.json"
TRACKER_SPREADSHEET = "Tracker -Candidates"
TRACKER_SHEET = "SCREENING"

os.makedirs(OUTPUT_DIR, exist_ok=True)
def out(name): return os.path.join(OUTPUT_DIR, name)

# =====================================================
# UTILITIES
# =====================================================
def color_nonkey_headers(xlsx_path, keep_cols=("name_location","designation","company")):
    wb = load_workbook(xlsx_path)
    ws = wb.active
    pink = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")
    for i, c in enumerate(ws[1], start=1):
        if c.value not in keep_cols:
            ws.cell(row=1, column=i).fill = pink
    wb.save(xlsx_path)

def load_list_from_excel(path):
    """Read Excel file into lowercase string set."""
    if not os.path.exists(path): return set()
    try:
        df = pd.read_excel(path, engine='openpyxl')
        col = next((c for c in df.columns if "designation" in c.lower()), None)
        if col:
            return set(df[col].dropna().astype(str).str.strip().str.lower())
    except Exception as e:
        print(f"⚠️ Could not read {path}: {e}")
    return set()

def append_to_excel(path, items):
    """Append unique lowercase items to Excel file."""
    if not items: return
    existing = load_list_from_excel(path)
    combined = sorted(set(existing).union([x.lower() for x in items]))
    pd.DataFrame({"designation": combined}).to_excel(path, index=False)
    print(f"✅ Updated {os.path.basename(path)} ({len(items)} new entries).")

# =====================================================
# PHASE 1 – CLEAN INPUT
# =====================================================
print("🔹 Phase 1: Cleaning input file...")
df = pd.read_excel(INPUT_FILE, engine='openpyxl')
df.columns = [c.strip().lower() for c in df.columns]

# --- find employment-detail column ---
employment_col = next((c for c in df.columns if "employment-detail" in c), None)
if not employment_col:
    raise Exception("❌ 'employment-detail' column not found.")

# --- split employment-detail into designation and company ---
def split_employment_detail(text):
    if pd.isna(text) or not isinstance(text, str):
        return "", ""
    parts = text.split(" at ", 1)
    if len(parts) == 2:
        return parts[0].strip(), parts[1].strip()
    else:
        return text.strip(), ""

designations, companies = zip(*[split_employment_detail(x) for x in df[employment_col]])

# --- insert new columns immediately after employment-detail ---
insert_index = df.columns.get_loc(employment_col) + 1
df.insert(insert_index, "designation", designations)
df.insert(insert_index + 1, "company", companies)

# --- add name_location column ---
df["name_location"] = (
    df["link"].fillna("").astype(str).str.strip() + "_" +
    df["location"].fillna("").astype(str).str.strip()
).str.replace("__", "_").str.strip("_")


# =====================================================
# LOAD MEMORY FILES
# =====================================================
selected_prev = load_list_from_excel(SELECTED_FILE)
removed_prev = load_list_from_excel(REMOVED_FILE)
print(f"✅ Loaded {len(selected_prev)} selected, {len(removed_prev)} removed designations.")

# =====================================================
# FILTER FOR GUI
# =====================================================
designations_all = sorted(set(df["designation"].dropna().astype(str).str.strip()))
designations_filtered = [d for d in designations_all if d.lower() not in removed_prev]
preselected = [d for d in designations_filtered if d.lower() in selected_prev]
new_designations = [d for d in designations_filtered if d.lower() not in selected_prev]

print(f"🧩 Total unique: {len(designations_all)} | Preselected: {len(preselected)} | New: {len(new_designations)}")

# =====================================================
# PYQT6 INTERFACE
# =====================================================
selected_final = set()
removed_final = set()

class Selector(QWidget):
    def __init__(self, all_desg, preselected):
        super().__init__()
        self.setWindowTitle("Finploy - Designation Selector")
        self.resize(700, 700)
        layout = QVBoxLayout()

        header = QLabel("Select Designations to Retain:")
        header.setStyleSheet("font-size:18px;font-weight:bold;color:#206A98;")
        layout.addWidget(header)

        self.search = QLineEdit()
        self.search.setPlaceholderText("Search designation...")
        self.search.textChanged.connect(self.filter_list)
        layout.addWidget(self.search)

        self.scroll = QScrollArea()
        self.frame = QFrame()
        self.vbox = QVBoxLayout(self.frame)

        self.checks = []
        for d in all_desg:
            cb = QCheckBox(d)
            cb.setChecked(d.lower() in [p.lower() for p in preselected])
            self.vbox.addWidget(cb)
            self.checks.append(cb)

        self.scroll.setWidget(self.frame)
        self.scroll.setWidgetResizable(True)
        layout.addWidget(self.scroll)

        btns = QHBoxLayout()
        submit = QPushButton("Submit & Continue")
        submit.setStyleSheet("background-color:#4EA647;color:white;font-weight:bold;")
        submit.clicked.connect(self.submit)
        btns.addWidget(submit)
        layout.addLayout(btns)
        self.setLayout(layout)

    def filter_list(self):
        term = self.search.text().lower()
        for cb in self.checks:
            cb.setVisible(term in cb.text().lower())

    def submit(self):
        for cb in self.checks:
            if cb.isChecked():
                selected_final.add(cb.text())
            else:
                removed_final.add(cb.text())
        QMessageBox.information(self, "Done",
                                f"✅ Selected: {len(selected_final)} | ❌ Removed: {len(removed_final)}")
        self.close()

# Launch GUI
app = QApplication(sys.argv)
win = Selector(designations_filtered, preselected)
win.show()
app.exec()

# =====================================================
# UPDATE MEMORY FILES
# =====================================================
append_to_excel(SELECTED_FILE, selected_final)
append_to_excel(REMOVED_FILE, removed_final)

# =====================================================
# FILTER FINAL DATA
# =====================================================
df["designation_lower"] = df["designation"].str.lower()
filtered = df[df["designation_lower"].isin(load_list_from_excel(SELECTED_FILE))]

phase1_path = out("phase1_output.xlsx")
filtered.to_excel(phase1_path, index=False)
color_nonkey_headers(phase1_path)
print(f"✅ phase1_output.xlsx saved ({len(filtered)} rows).")

# =====================================================
# COMPARE WITH TRACKER (FIXED)
# =====================================================
print("🔹 Phase 2.4: Comparing with Tracker...")
scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
credentials = ServiceAccountCredentials.from_json_keyfile_name(SERVICE_ACCOUNT_JSON, scope)
client = gspread.authorize(credentials)

sheet = client.open(TRACKER_SPREADSHEET).worksheet(TRACKER_SHEET)
values = sheet.get_all_values()
tracker_df = pd.DataFrame(values[1:], columns=values[0])
tracker_df.columns = [c.strip().lower() for c in tracker_df.columns]

# ⭐⭐⭐ FIX ADDED HERE — normalize both sides ⭐⭐⭐
filtered["name_location"] = filtered["name_location"].astype(str).str.strip()
tracker_df["name_location"] = tracker_df["name_location"].astype(str).str.strip()

# For even safer matching (case-insensitive)
filtered["clean_key"] = filtered["name_location"].str.lower()
tracker_df["clean_key"] = tracker_df["name_location"].str.lower()

# Perform comparison
unscreened_df = filtered[~filtered["clean_key"].isin(tracker_df["clean_key"])].copy()
screened_df = filtered[filtered["clean_key"].isin(tracker_df["clean_key"])].copy()

unscreened_path = out("output1.xlsx")
unscreened_df.to_excel(unscreened_path, index=False)
screened_path = out("screened_candidates.xlsx")
screened_df.to_excel(screened_path, index=False)

print(f"🎉 Complete | UNSCREENED: {len(unscreened_df)} | SCREENED: {len(screened_df)}")

# =====================================================
# RUN NEXT FILE
# =====================================================
try:
    print("▶️ Running main9.py ...")
    subprocess.run(["python", r"D:\matching_harsh\Job_matching_unscreened\main9.py"], check=True)
    print("✅ main9.py executed successfully!")
except Exception as e:
    print(f"❌ Failed to run main9.py: {e}")
